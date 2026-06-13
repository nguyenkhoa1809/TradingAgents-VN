"""
Thesis Tracker — parse TradingAgents HTML reports → Excel
Zero API calls. Run after main.py to get a consolidated view.

Output: reports/_tracker/thesis_tracker.xlsx
  Sheet "Latest"   — most recent signal per ticker
  Sheet "History"  — all runs, chronological
"""

import re
import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


REPORTS_DIR = Path(__file__).parent / "reports"
OUT_DIR     = REPORTS_DIR / "_tracker"
OUT_PATH    = OUT_DIR / "thesis_tracker.xlsx"

RATING_MAP = {
    "overweight": "BUY",
    "buy":        "BUY",
    "underweight":"SELL",
    "sell":       "SELL",
    "neutral":    "HOLD",
    "hold":       "HOLD",
    "equalweight":"HOLD",
    "equal-weight":"HOLD",
}

SIGNAL_COLOR = {
    "BUY":  "1a472a",   # dark green
    "SELL": "4a1020",   # dark red
    "HOLD": "2d3a1a",   # dark olive
}
SIGNAL_FG = {
    "BUY":  "34d399",
    "SELL": "f87171",
    "HOLD": "fbbf24",
}


# ── Parser ────────────────────────────────────────────────────────────────────

def parse_report(html_path: Path) -> dict | None:
    """Extract key fields from a single TradingAgents HTML report."""
    # Filename: TICKER_DATE_PROVIDER_vN.html
    parts = html_path.stem.split("_")
    if len(parts) < 4:
        return None
    ticker   = parts[0]
    date_str = parts[1]          # YYYY-MM-DD
    provider = parts[2]
    version  = parts[3]          # vN

    try:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return None

    content = html_path.read_text(encoding="utf-8", errors="replace")

    # ── Isolate final_trade_decision section (Portfolio Manager only) ──
    pm_section = ""
    idx_pm = content.find('id="final_trade_decision"')
    if idx_pm != -1:
        # Take up to next section-card or end of main
        end_idx = content.find('<div class="section-card"', idx_pm + 10)
        pm_section = content[idx_pm: end_idx if end_idx != -1 else idx_pm + 8000]

    search_scope = pm_section if pm_section else content

    # ── Rating / signal — from PM section only ──
    rating_raw = ""
    # Pattern 1: <strong>Rating</strong>: Overweight
    m = re.search(r"<strong>Rating</strong>\s*:\s*([^\n<]{1,30})", search_scope, re.IGNORECASE)
    if m:
        rating_raw = m.group(1).strip().rstrip(".")
    # Pattern 2: "Final Trading Decision: <strong>HOLD</strong>" or plain text
    if not rating_raw:
        m = re.search(
            r"Final Trading Decision\s*:\s*(?:<[^>]+>)?\s*(BUY|SELL|HOLD|OVERWEIGHT|UNDERWEIGHT|NEUTRAL)",
            search_scope, re.IGNORECASE
        )
        if m:
            rating_raw = m.group(1).strip()
    # Pattern 3: signal-value class in full document (header banner)
    if not rating_raw:
        m = re.search(
            r'class=["\']signal-value["\'][^>]*>\s*(BUY|SELL|HOLD|OVERWEIGHT|UNDERWEIGHT|NEUTRAL)',
            content, re.IGNORECASE
        )
        if m:
            rating_raw = m.group(1).strip()

    signal = RATING_MAP.get(rating_raw.lower(), rating_raw.upper()[:4] if rating_raw else "?")

    # ── Stop loss — PM section first, then full doc ──
    stop_loss = ""
    for scope in (search_scope, content):
        m = re.search(r"<strong>Stop[\s\-]Loss[^<]*</strong>\s*:\s*([\d,\.]+)", scope, re.IGNORECASE)
        if m:
            stop_loss = m.group(1).replace(",", "")
            break
    if not stop_loss:
        m2 = re.search(r"stop[\s\-]loss\s+(?:at|of|@)\s+([\d,\.]+)", search_scope, re.IGNORECASE)
        if m2:
            stop_loss = m2.group(1).replace(",", "")

    # ── Price target — PM section ──
    price_target = ""
    m = re.search(r"<strong>Price[\s\-]?Target[^<]*</strong>\s*:\s*([\d,\.]+)", search_scope, re.IGNORECASE)
    if m:
        price_target = m.group(1).replace(",", "")

    # ── Entry price — PM section ──
    entry_price = ""
    m = re.search(r"<strong>Entry[\s\-]?Price[^<]*</strong>\s*:\s*([\d,\.]+)", search_scope, re.IGNORECASE)
    if m:
        entry_price = m.group(1).replace(",", "")

    # ── Executive Summary — PM section ──
    summary = ""
    m = re.search(
        r"<strong>Executive Summary</strong>\s*:\s*(.*?)(?=<p>|</div>|$)",
        search_scope, re.IGNORECASE | re.DOTALL
    )
    if m:
        raw = re.sub(r"<[^>]+>", "", m.group(1))
        summary = " ".join(raw.split())[:350]
    if not summary and pm_section:
        paras = re.findall(r"<p>(.*?)</p>", pm_section, re.DOTALL)
        if paras:
            raw = re.sub(r"<[^>]+>", "", paras[0])
            summary = " ".join(raw.split())[:350]

    # ── Position sizing — PM section ──
    position_size = ""
    m = re.search(r"<strong>Position Siz[^<]*</strong>\s*:\s*(.*?)(?=</p>|<p>)", search_scope, re.IGNORECASE | re.DOTALL)
    if m:
        raw = re.sub(r"<[^>]+>", "", m.group(1))
        position_size = " ".join(raw.split())[:150]

    return {
        "ticker":        ticker,
        "date":          date,
        "provider":      provider,
        "version":       version,
        "signal":        signal,
        "rating_raw":    rating_raw,
        "stop_loss":     stop_loss,
        "price_target":  price_target,
        "entry_price":   entry_price,
        "position_size": position_size,
        "summary":       summary,
        "file":          html_path.name,
    }


# ── Excel writer ──────────────────────────────────────────────────────────────

def col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def header_row(ws, headers, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font      = Font(bold=True, color="F1F5F9", size=10)
        cell.fill      = PatternFill("solid", fgColor="0F172A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="1F2D45")
        cell.border = Border(bottom=thin)


def signal_cell(ws, row, col, signal):
    cell = ws.cell(row=row, column=col, value=signal)
    bg   = SIGNAL_COLOR.get(signal, "1e293b")
    fg   = SIGNAL_FG.get(signal, "94a3b8")
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color=fg, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    return cell


def write_sheet(ws, rows, is_latest=True):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0F172A"

    if is_latest:
        headers = ["Ticker", "Date", "Provider", "Signal", "Entry", "Stop Loss",
                   "Price Target", "Position Size", "Executive Summary"]
        col_widths = [8, 12, 10, 8, 12, 12, 12, 28, 80]
    else:
        headers = ["Ticker", "Date", "Provider", "Ver", "Signal",
                   "Entry", "Stop Loss", "Price Target", "Summary"]
        col_widths = [8, 12, 10, 5, 8, 12, 12, 12, 80]

    header_row(ws, headers)

    for c, w in enumerate(col_widths, 1):
        col_width(ws, c, w)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    thin = Side(style="thin", color="1F2D45")

    for r_idx, row in enumerate(rows, 2):
        ws.row_dimensions[r_idx].height = 40

        if is_latest:
            vals = [
                row["ticker"], str(row["date"]), row["provider"],
                row["signal"], row["entry_price"], row["stop_loss"],
                row["price_target"], row["position_size"], row["summary"],
            ]
        else:
            vals = [
                row["ticker"], str(row["date"]), row["provider"], row["version"],
                row["signal"], row["entry_price"], row["stop_loss"],
                row["price_target"], row["summary"],
            ]

        for c_idx, val in enumerate(vals, 1):
            if is_latest and c_idx == 4 or not is_latest and c_idx == 5:
                signal_cell(ws, r_idx, c_idx, val)
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font      = Font(color="CBD5E1", size=9)
                cell.alignment = Alignment(
                    vertical="center", wrap_text=(c_idx == len(vals))
                )
                cell.fill = PatternFill("solid", fgColor="0A0F1E" if r_idx % 2 == 0 else "0F1629")
                cell.border = Border(bottom=Side(style="thin", color="1A2540"))

    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Collect all reports
    all_records = []
    for html_file in sorted(REPORTS_DIR.rglob("*.html")):
        if html_file.parent.name == "_tracker":
            continue
        rec = parse_report(html_file)
        if rec and rec["signal"]:
            all_records.append(rec)

    if not all_records:
        print("No reports found.")
        return

    # Sort by ticker then date
    all_records.sort(key=lambda r: (r["ticker"], r["date"], r["version"]))
    print(f"Parsed {len(all_records)} reports across "
          f"{len({r['ticker'] for r in all_records})} tickers")

    # Latest: one row per ticker (most recent date + highest version)
    latest_map: dict[str, dict] = {}
    for rec in all_records:
        key = rec["ticker"]
        if key not in latest_map:
            latest_map[key] = rec
        else:
            prev = latest_map[key]
            if (rec["date"], rec["version"]) > (prev["date"], prev["version"]):
                latest_map[key] = rec

    latest_rows = sorted(latest_map.values(), key=lambda r: r["ticker"])

    # ── Write workbook ──
    wb = openpyxl.Workbook()

    # Sheet 1 — Latest
    ws_latest = wb.active
    ws_latest.title = "Latest Signals"
    write_sheet(ws_latest, latest_rows, is_latest=True)

    # Sheet 2 — History
    ws_hist = wb.create_sheet("History")
    write_sheet(ws_hist, all_records, is_latest=False)

    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH.resolve()}")

    # Summary
    buys  = [r["ticker"] for r in latest_rows if r["signal"] == "BUY"]
    holds = [r["ticker"] for r in latest_rows if r["signal"] == "HOLD"]
    sells = [r["ticker"] for r in latest_rows if r["signal"] == "SELL"]
    print(f"\n📊 Latest signals ({len(latest_rows)} tickers):")
    print(f"  🟢 BUY  ({len(buys)}):  {', '.join(buys)}")
    print(f"  🟡 HOLD ({len(holds)}): {', '.join(holds)}")
    print(f"  🔴 SELL ({len(sells)}): {', '.join(sells)}")


if __name__ == "__main__":
    main()
